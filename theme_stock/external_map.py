"""L0 外部产业链图谱导入器 — 解析 xlsx → 差异审计 → 入库

用法:
  python -m theme_stock.external_map --dry-run   # 仅解析+差异报告, 不入库
  python -m theme_stock.external_map              # 解析+差异报告+入库
"""

from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import warnings
import openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

sys.path.insert(0, str(Path(__file__).parent.parent))

from theme_stock.store import ThemeStockStore

MAP_DIR = Path(__file__).parent.parent / "alpha产业图谱"
OVERLAP_THRESHOLD = 0.6   # 较小集合中至少60%出现在较大集合中
MIN_OVERLAP_ABS = 5        # 至少5只重合标的才认为匹配


# ============================================================
# Parse
# ============================================================

def _forward_fill(rows: list[list]) -> list[list]:
    """对层级1/层级2列做 forward-fill 处理合并单元格"""
    if not rows:
        return rows
    for col_idx in (1, 2):  # 层级1, 层级2
        carry = ""
        for row in rows:
            val = (row[col_idx] or "").strip() if col_idx < len(row) else ""
            if val:
                carry = val
            else:
                row[col_idx] = carry
    return rows


def parse_xlsx(filepath: Path) -> dict:
    """解析单个 xlsx, 返回 {industry, tiers: {tier: {segment: [(code, name, desc), ...]}}}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_rows.append(list(row))

    rows = _forward_fill(raw_rows)

    industry_name = _extract_industry_from_filename(filepath.name)

    tiers: dict[str, dict[str, list[tuple[str, str, str]]]] = defaultdict(lambda: defaultdict(list))

    for row in rows:
        tier = (row[1] or "").strip()
        segment = (row[2] or "").strip()
        raw_code = (row[3] or "").strip()
        name = (row[4] or "").strip()
        desc = (row[10] or "").strip() if len(row) > 10 else ""

        code = _normalize_code(raw_code)
        if not code or not name:
            continue
        if not tier or not segment:
            continue
        if segment == "-":
            segment = tier  # 部分 xlsx 层级2 为空时用 "-" 占位

        tiers[tier][segment].append((code, name, desc))

    wb.close()
    return {"industry": industry_name, "tiers": dict(tiers)}


def _extract_industry_from_filename(filename: str) -> str:
    """从文件名提取产业名: 'AI算力产业链20260622涨跌幅.xlsx' → 'AI算力'"""
    name = filename.replace(".xlsx", "")
    name = re.sub(r"\d{8}.*$", "", name)
    name = _normalize_industry(name)
    return name


def _normalize_industry(name: str) -> str:
    """产业名规范化: 去常见后缀"""
    for suffix in ["产业链", "概念", "梳理", "产业链图谱", "产业图谱"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    return name.strip()


def _normalize_code(raw: str) -> str:
    """股票代码标准化: '000034.SZ' → '000034', '蔚来' → '' """
    raw = raw.strip()
    if not raw:
        return ""
    m = re.match(r"(\d{6})", raw)
    if m:
        return m.group(1)
    return ""


def parse_all(dirpath: Path | None = None) -> dict[str, dict]:
    """解析全部 xlsx, 返回 {产业名: parsed_data}"""
    if dirpath is None:
        dirpath = MAP_DIR
    all_data: dict[str, dict] = {}
    files = sorted(dirpath.glob("*.xlsx"))
    for fp in files:
        try:
            data = parse_xlsx(fp)
            if data["tiers"]:
                all_data[data["industry"]] = data
        except Exception as e:
            print(f"  [WARN] 解析失败 {fp.name}: {e}")
    return all_data


# ============================================================
# Match
# ============================================================

def _overlap_coef(set_a: set, set_b: set) -> float:
    """Overlap Coefficient: |A∩B| / min(|A|, |B|)"""
    if not set_a or not set_b:
        return 0.0
    return len(set_a & set_b) / min(len(set_a), len(set_b))


def match_to_existing(parsed: dict, store: ThemeStockStore) -> dict[str, tuple[str | None, float, int]]:
    """对每个解析出的产业, 与 DB 中已有产业做标的重合度匹配

    使用 Overlap Coefficient (更适合大小集合非对称的场景: DB精选龙头 vs xlsx全覆盖)
    Returns: {parsed_name: (matched_canonical_or_None, overlap_coef, overlap_abs)}
    """
    db_industries = _get_db_industry_stocks(store)

    results = {}
    for pname, pdata in parsed.items():
        p_stocks = _extract_stock_set(pdata)
        best_match = None
        best_coef = 0.0
        best_overlap = 0

        for dname, d_stocks in db_industries.items():
            overlap = len(p_stocks & d_stocks)
            coef = _overlap_coef(p_stocks, d_stocks)
            if coef > best_coef:
                best_coef = coef
                best_match = dname
                best_overlap = overlap
            elif coef == best_coef and overlap > best_overlap:
                best_match = dname
                best_overlap = overlap

        if best_coef >= OVERLAP_THRESHOLD and best_overlap >= MIN_OVERLAP_ABS:
            results[pname] = (best_match, best_coef, best_overlap)
        else:
            results[pname] = (None, best_coef, 0)

    return results


def _get_db_industry_stocks(store: ThemeStockStore) -> dict[str, set[str]]:
    """获取 DB 中每个产业的全部标的代码集合"""
    conn = store._get_conn()
    cur = conn.execute("SELECT DISTINCT industry, code FROM chain_map")
    result: dict[str, set[str]] = defaultdict(set)
    for row in cur:
        result[row["industry"]].add(row["code"])
    return dict(result)


def _extract_stock_set(parsed_data: dict) -> set[str]:
    """从解析数据中提取全部标的代码"""
    codes = set()
    for tier_segments in parsed_data["tiers"].values():
        for segments in tier_segments.values():
            for code, _, _ in segments:
                codes.add(code)
    return codes


# ============================================================
# Diff
# ============================================================

def generate_diff(parsed: dict, matches: dict, store: ThemeStockStore) -> str:
    """生成差异报告"""
    lines = []
    lines.append("=" * 72)
    lines.append("L0 外部图谱差异报告")
    lines.append("=" * 72)

    new_industries = []
    matched_industries = []
    total_new_stocks = 0
    total_existing_stocks = 0
    total_conflict_segments = 0

    for pname, pdata in parsed.items():
        matched_canonical, coef, overlap = matches[pname]
        p_stocks_count = len(_extract_stock_set(pdata))

        if matched_canonical is None:
            new_industries.append((pname, p_stocks_count))
            total_new_stocks += p_stocks_count
        else:
            db_stocks_count = len(_get_db_industry_stocks(store).get(matched_canonical, set()))
            conflicts = _find_conflicts(pname, pdata, matched_canonical, store)
            matched_industries.append((pname, matched_canonical, coef, overlap, p_stocks_count, db_stocks_count, conflicts))
            total_existing_stocks += p_stocks_count
            total_conflict_segments += len(conflicts)

    # 汇总
    lines.append(f"\nSummary: {len(parsed)} industries, "
                 f"NEW {len(new_industries)}, "
                 f"MATCHED {len(matched_industries)}")
    lines.append(f"   ~{total_new_stocks} stocks in new industries, "
                 f"~{total_existing_stocks} stocks added to existing")
    if total_conflict_segments:
        lines.append(f"   [!] {total_conflict_segments} segments with stock differences (BOM-priority preserved)")

    # 全新产业
    if new_industries:
        lines.append(f"\n{'─' * 72}")
        lines.append(f"[NEW] Industries ({len(new_industries)})")
        lines.append(f"{'─' * 72}")
        for name, count in sorted(new_industries, key=lambda x: -x[1]):
            tiers_count = len(parsed[name]["tiers"])
            lines.append(f"  {name}  ({tiers_count} tiers, {count} stocks)")

    # 匹配产业
    if matched_industries:
        lines.append(f"\n{'─' * 72}")
        lines.append(f"[MATCHED] Industries ({len(matched_industries)})")
        lines.append(f"{'─' * 72}")
        for pname, canonical, coef, overlap, p_count, db_count, conflicts in sorted(
            matched_industries, key=lambda x: -x[2]
        ):
            lines.append(f"\n  {pname} -> {canonical}  (Overlap={coef:.2f}, common={overlap})")
            lines.append(f"    xlsx new: {p_count - overlap} | DB existing: {db_count}")
            if conflicts:
                lines.append(f"    [!] Conflict segments ({len(conflicts)}) - BOM kept, xlsx appended:")
                for seg_name, bom_only, xlsx_only in conflicts[:5]:
                    lines.append(f"      [{seg_name}] BOM-only: {bom_only} | xlsx-only: {xlsx_only}")
                if len(conflicts) > 5:
                    lines.append(f"      ... {len(conflicts)} total conflict segments")

    lines.append(f"\n{'─' * 72}")
    lines.append("Strategy: BOM priority - existing (industry,tier,segment,code) kept; xlsx fills gaps")
    lines.append("=" * 72)

    return "\n".join(lines)


def _find_conflicts(pname: str, pdata: dict, db_industry: str, store: ThemeStockStore) -> list[tuple[str, list[str], list[str]]]:
    """找出同一 segment 下 BOM 和 xlsx 标的列表的差异"""
    db_segments = _get_db_segment_stocks(store, db_industry)
    conflicts = []

    for tier, segments in pdata["tiers"].items():
        for seg, stocks in segments.items():
            xlsx_codes = {c for c, _, _ in stocks}
            db_codes = db_segments.get((tier, seg), set())

            bom_only = sorted(db_codes - xlsx_codes)
            xlsx_only = sorted(xlsx_codes - db_codes)

            if bom_only or xlsx_only:
                seg_label = f"{tier}/{seg}" if tier != seg else seg
                conflicts.append((seg_label, bom_only, xlsx_only))

    return conflicts


def _get_db_segment_stocks(store: ThemeStockStore, industry: str) -> dict[tuple[str, str], set[str]]:
    """获取 DB 中某个产业的 (tier, segment) → codes 映射"""
    conn = store._get_conn()
    cur = conn.execute(
        "SELECT tier, segment, code FROM chain_map WHERE industry=?",
        (industry,),
    )
    result: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in cur:
        result[(row["tier"], row["segment"])].add(row["code"])
    return dict(result)


# ============================================================
# Import
# ============================================================

def import_to_db(parsed: dict, matches: dict, store: ThemeStockStore) -> dict:
    """将解析数据导入 chain_map, BOM 优先

    Returns: {new: N, skipped: N, total_industries: N}
    """
    new_count = 0
    skipped_count = 0
    existing_codes = _load_existing_codes(store)

    for pname, pdata in parsed.items():
        matched_canonical, _, _ = matches[pname]
        industry = matched_canonical if matched_canonical else pname

        batch = []
        for tier, segments in pdata["tiers"].items():
            for seg, stocks in segments.items():
                segment = seg if seg != "-" else tier
                for code, name, desc in stocks:
                    key = (industry, tier, segment, code)
                    if key in existing_codes:
                        skipped_count += 1
                        continue

                    batch.append({
                        "industry": industry,
                        "tier": tier,
                        "segment": segment,
                        "code": code,
                        "name": name,
                        "market": "A",
                        "role": desc[:200] if desc else "",
                        "source": "external_map",
                        "source_ver": "alpha_20260622",
                        "confidence": "medium",
                        "is_verified": 0,
                    })
                    new_count += 1
                    existing_codes.add(key)

        if batch:
            store.upsert_chain_batch(batch)

    # 写入 alias 映射
    for pname, (matched, _, _) in matches.items():
        if matched and matched != pname:
            store.add_alias(pname, matched, "external_map_auto")

    stats = store.get_chain_stats()
    return {"new": new_count, "skipped": skipped_count,
            "industries": stats["industries"], "stocks": stats["stocks"]}


def _load_existing_codes(store: ThemeStockStore) -> set[tuple]:
    conn = store._get_conn()
    cur = conn.execute("SELECT industry, tier, segment, code FROM chain_map")
    return {(r["industry"], r["tier"], r["segment"], r["code"]) for r in cur}


# ============================================================
# Main
# ============================================================

def run(dry_run: bool = False):
    store = ThemeStockStore()
    store.init_db()

    print("[1/4] Parsing xlsx files...")
    parsed = parse_all()
    print(f"      Parsed: {len(parsed)} industries")

    if not parsed:
        print("      [ERROR] No xlsx files found")
        return

    print("[2/4] Matching to existing DB industries...")
    matches = match_to_existing(parsed, store)

    new_count = sum(1 for m in matches.values() if m[0] is None)
    matched_count = sum(1 for m in matches.values() if m[0] is not None)
    print(f"      Matched: {matched_count} -> existing | {new_count} -> new")

    print("[3/4] Diff report:")
    report = generate_diff(parsed, matches, store)
    print(report)

    if dry_run:
        print("\n[DRY RUN] No data imported. Run without --dry-run to commit.")
        store.close()
        return

    print("[4/4] Importing (BOM priority)...")
    result = import_to_db(parsed, matches, store)
    print(f"      New: {result['new']} rows | Skipped: {result['skipped']} rows (BOM/SC already present)")
    print(f"      DB now: {result['industries']} industries, {result['stocks']} stocks")

    store.close()
    print("Done.")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    run(dry_run=dry)
