"""每日仪表盘 — 五维状态+采集管线+可操作警报+趋势。

在 daily_collect 末尾调用，产出 reports/Dashboard.md 并同步到根目录。
Obsidian 中可用 Homepage 插件设为首页，或固定标签页。
"""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path

import store
from config import REPORT_DIR
from data import extract_codes_from_text
from engine_sector_rotation import sector_frequency, sector_persistence, sector_stocks
from dual_track_screener import screen as dual_track_screen
from engine_theme_lifecycle import lifecycle as theme_lifecycle
from engine_market_rhythm import classify_rhythm_stage, daily_rhythm
from engine_similar_days import similar_sectors
from engine_leader_backtest import leader_frequency


def _last_trading_day(ref: date | None = None) -> date:
    """最近一个交易日（跳过周末）。"""
    d = (ref or date.today()) - timedelta(days=1)
    while d.weekday() >= 5:
        d = d - timedelta(days=1)
    return d

DASHBOARD_PATH = REPORT_DIR / "Dashboard.md"

# === 五维 → collector 映射 ===
DIMENSIONS = [
    {
        "name": "① 公告深研",
        "icon": "📄",
        "collector": "announcement_deep_read",
        "data_source": "announcements",
        "cost": "$9/天",
        "desc": "全市场公告 → 四层筛选 → LLM五维评分 → Obsidian存档",
        "signal_query": "announce_signal",
    },
    {
        "name": "② 研报跟踪",
        "icon": "📊",
        "collector": "research_deep_read",
        "data_source": "research_reports",
        "cost": "$1/天",
        "desc": "东方财富全市场API → 信号检测 → LLM分析 → 累积档案",
        "signal_query": "dossier",
    },
    {
        "name": "③ 调研情绪",
        "icon": "🔍",
        "collector": "sentiment_track",
        "data_source": "inst_survey",
        "cost": "$0",
        "desc": "机构调研 → 密集/首次/升温检测 → 追加到档案",
        "signal_query": "dossier",
    },
    {
        "name": "④ 互动易",
        "icon": "💬",
        "collector": "interactions",
        "data_source": "interactions",
        "cost": "$0",
        "desc": "深交所+上交所互动易 → 关键词检测 → 追加到档案",
        "signal_query": "dossier",
    },
    {
        "name": "⑤ 业绩预告",
        "icon": "📈",
        "collector": "earnings",
        "data_source": "earnings_forecast",
        "cost": "$0",
        "desc": "全市场业绩预告/快报 → 暴增/扭亏/暴雷检测 → 追加到档案",
        "signal_query": "none",
    },
]

# === 采集管线中文名 ===
SOURCE_LABELS = {
    "announcements": "📄 公告采集",
    "announcement_deep_read": "📄 公告深研",
    "research": "📊 研报采集",
    "research_deep_read": "📊 研报跟踪",
    "surveys": "🔍 机构调研",
    "sentiment_track": "🔍 调研+互动情绪",
    "interactions": "💬 互动易",
    "earnings": "📈 业绩预告",
    "news": "📰 个股新闻",
    "news_signals": "📰 新闻边际信号",
    "industry": "🏭 行业研报",
    "industry_deep_read": "🏭 行业深度分析",
    "catalyst_tracker": "📡 催化走势跟踪",
    "wechat": "💚 微信公众号",
    "weibo": "🐦 唐史主任微博",
    "zsxq": "⭐ 知识星球",
    "jiuyang": "📝 韭研脱水研报",
    "lockups": "🔒 限售解禁",
    "eps": "📐 一致预期EPS",
    "financials": "💰 财务指标",
}

# === 可操作警报：异常 → 排查指引 ===
ALERT_ACTIONS = {
    "announcement_deep_read": {
        "超时": "检查 DeepSeek API 是否正常 → 减少 --days 范围 → 或等下次重试",
        "无公告数据": "检查 akshare.stock_notice_report 接口 → 确认交易日历",
    },
    "research": {
        "从未运行": "手动执行 python daily_collect.py --source research 验证 API",
    },
    "research_deep_read": {
        "从未运行": "手动执行 python daily_collect.py --source research_deep_read",
    },
    "wechat": {
        "RSS数据陈旧": "打开 http://111.231.44.12:4000/dash 检查 WeWe-RSS 是否需要重新登录",
    },
    "interactions": {
        "超时": "日频72只逐只调约需6分钟 → 检查网络 → 或降为周频",
    },
    "zsxq": {
        "从未运行": "检查知识星球 cookie 是否过期 → 需手动更新",
    },
}


def _status_icon(status: str) -> str:
    return {"ok": "✅", "error": "❌", "timeout": "⏰", "skip": "➖"}.get(status, "⚠️")


def _collector_status() -> list[dict]:
    rows = []
    try:
        with store._conn() as conn:
            for src in SOURCE_LABELS:
                r = conn.execute(
                    "SELECT * FROM collect_status WHERE source = ? ORDER BY last_run_at DESC LIMIT 1",
                    (src,),
                ).fetchone()
                row = dict(r) if r else {
                    "source": src, "status": "unknown", "last_date": "",
                    "last_run_at": "", "message": "从未运行", "added_count": 0,
                }
                # 提取 HH:MM
                ts = row.get("last_run_at", "") or ""
                row["run_time"] = ts[-5:] if len(ts) >= 16 and ts[10] == " " else ""
                rows.append(row)
    except Exception:
        pass
    return rows


def _deep_read_weekly() -> dict:
    try:
        with store._conn() as conn:
            rows = conn.execute(
                "SELECT date, COUNT(*) as cnt, SUM(CASE WHEN total_score>=60 THEN 1 ELSE 0 END) as a60 "
                "FROM deep_read_results WHERE date >= ? GROUP BY date ORDER BY date",
                ((date.today() - timedelta(days=7)).isoformat(),),
            ).fetchall()
        daily = {r["date"]: (r["cnt"] or 0, r["a60"] or 0) for r in rows}
        total_a60 = sum(v[1] for v in daily.values())
        return {"daily": daily, "total_a60": total_a60}
    except Exception:
        return {"daily": {}, "total_a60": 0}


def _dossier_count() -> int:
    d = REPORT_DIR / "research_dossiers"
    return len(list(d.glob("*.md"))) if d.exists() else 0


def _score_dossier(text: str, latest_date: str) -> tuple[int, str]:
    """计算研报档案星级评分 (1-5).

    维度: 信号强度(30) + LLM评分(30) + 机构数(15) + 信号频率(15) + 新鲜度(10)
    """
    import re
    score = 0

    # 1. 信号强度 (绝对值加总, max 30)
    weights = re.findall(r"\(([+-]?\d+)分\)", text)
    signal_sum = sum(abs(int(w)) for w in weights)
    score += min(signal_sum * 2, 30)

    # 2. LLM评分 (max 30)
    llm_m = re.search(r"评分:\s*(\d+)", text)
    if llm_m:
        score += min(int(llm_m.group(1)) * 0.3, 30)

    # 3. 机构覆盖 (评级表行数, max 15)
    inst_count = len(re.findall(r"^\| \d{4}-\d{2}-\d{2} \|", text, re.MULTILINE))
    score += min(inst_count * 3, 15)

    # 4. 信号频次 (不同日期数, max 15)
    sig_dates = set(re.findall(r"### (\d{4}-\d{2}-\d{2})", text))
    sig_dates |= set(re.findall(r"## .+ \((\d{4}-\d{2}-\d{2})\)", text))
    score += min(len(sig_dates) * 5, 15)

    # 5. 新鲜度 (max 10)
    if latest_date:
        try:
            days = (date.today() - date.fromisoformat(latest_date)).days
            if days <= 1: score += 10
            elif days <= 3: score += 6
            elif days <= 7: score += 3
            elif days <= 14: score += 1
        except Exception:
            pass

    STARS = {5: "★★★★★", 4: "★★★★", 3: "★★★", 2: "★★", 1: "★"}
    if score >= 70: n = 5
    elif score >= 50: n = 4
    elif score >= 30: n = 3
    elif score >= 15: n = 2
    else: n = 1
    return (score, f'<span class="star-{n}">{STARS[n]}</span>', STARS[n])


def _generate_dossier_index():
    """生成研报档案索引 README → reports/research_dossiers/README.md"""
    import re
    d = REPORT_DIR / "research_dossiers"
    today = date.today()
    week_ago = today - timedelta(days=7)

    dossiers = []
    for fp in sorted(d.glob("*.md")):
        if fp.name == "README.md":
            continue
        try:
            text = fp.read_text(encoding="utf-8")
        except Exception:
            continue
        dates = re.findall(r"### (\d{4}-\d{2}-\d{2})", text)
        dates += re.findall(r"## .+ \((\d{4}-\d{2}-\d{2})\)", text)
        latest = max(dates) if dates else fp.stat().st_mtime
        if isinstance(latest, float):
            from datetime import datetime
            latest = datetime.fromtimestamp(latest).strftime("%Y-%m-%d")
        nm = re.search(r'^name:\s*"?(.+?)"?\s*$', text, re.MULTILINE)
        name = nm.group(1).strip() if nm else ""
        display = f"{name} ({fp.stem[:6]})" if name else fp.stem[:6]
        sigs = re.findall(r"- \[(\w+)\]\s*(.+?)(?:\s*\([+-]?\d+分\))?\s*$", text, re.MULTILINE)
        sig_type, sig_desc = sigs[-1] if sigs else ("", "")
        raw_score, html_stars, plain_stars = _score_dossier(text, latest)
        dossiers.append({
            "display": display, "code": fp.stem[:6],
            "latest_date": latest,
            "sig_type": sig_type, "sig_desc": sig_desc,
            "stars": html_stars, "score": raw_score,
        })

    # 按日期倒序, 同日期按星级降序
    dossiers.sort(key=lambda x: (x["latest_date"], x["score"]), reverse=True)

    today_list = [x for x in dossiers if x["latest_date"] == today.isoformat()]
    week_list = [x for x in dossiers if today.isoformat() > x["latest_date"] >= week_ago.isoformat()]
    older_list = [x for x in dossiers if x["latest_date"] < week_ago.isoformat()]

    buf = [
        f"# 研报档案索引",
        "",
        f"> 自动生成于 {today.isoformat()} | 共 {len(dossiers)} 份档案",
        "",
    ]

    sections = [
        (f"## 🔥 今日更新 ({len(today_list)})", today_list),
        (f"## 📅 近7天更新 ({len(week_list)})", week_list),
        (f"## 📦 更早 ({len(older_list)})", older_list),
    ]
    for heading, items in sections:
        buf.append(heading)
        buf.append("")
        if items:
            buf.append("| 评级 | 标的 | 代码 | 最后信号日 | 最新信号 |")
            buf.append("|:----:|------|:----:|:----------:|----------|")
            for x in items:
                sig = f"{x['sig_type']}: {x['sig_desc'][:30]}" if x["sig_type"] else "—"
                buf.append(f"| {x['stars']} | {x['display']} | {x['code']} | {x['latest_date']} | {sig} |")
        else:
            buf.append("_暂无_")
        buf.append("")

    d.mkdir(parents=True, exist_ok=True)
    (d / "README.md").write_text("\n".join(buf), encoding="utf-8")


def _data_freshness(source: str) -> tuple[str, str]:
    """(新鲜度文字, 天数)"""
    col_map = {
        "announcements": "date", "research_reports": "report_date",
        "inst_survey": "notice_date", "interactions": "reply_time",
        "earnings_forecast": "notice_date",
    }
    if source not in col_map:
        return "—", ""
    try:
        with store._conn() as conn:
            col = col_map[source]
            r = conn.execute(
                f"SELECT MAX({col}) FROM {source} LIMIT 1"
            ).fetchone()[0]
            if r:
                days = (date.today() - date.fromisoformat(str(r)[:10])).days
                if days <= 1:
                    return "🟢 新鲜", str(days)
                elif days <= 3:
                    return f"🟡 {days}天前", str(days)
                else:
                    return f"🔴 {days}天前", str(days)
    except Exception:
        pass
    return "—", ""


def _engine_status() -> list[dict]:
    """检查分析引擎产出文件，提取关键数字。"""
    today = date.today()
    engines = []

    def _mtime(path: Path) -> str:
        try:
            ts = datetime.fromtimestamp(path.stat().st_mtime)
            return ts.strftime("%m-%d %H:%M")
        except Exception:
            return ""

    # catalyst_screen
    cs_path = REPORT_DIR / "catalyst" / f"catalyst_screen_{today.isoformat()}.md"
    cs_json = REPORT_DIR / "catalyst" / f"catalyst_screen_{today.isoformat()}.json"
    cs_ok = cs_path.exists()
    cs_nums = ""
    if cs_ok:
        try:
            import json, re
            data = json.loads(cs_json.read_text(encoding="utf-8"))
            cats = data.get("catalysts", [])
            high_n = sum(1 for c in cats if c.get("final_actionability", 0) >= 40)
            crit_n = sum(1 for c in cats if c.get("final_actionability", 0) >= 60)
            cs_nums = f"{len(cats)}催化, {high_n} HIGH+, {crit_n} CRITICAL"
            # 映射质量审计
            audit = data.get("stock_map_audit", {})
            if audit:
                total_m = audit.get("total_mappings", 0)
                llm_pct = audit.get("llm_direct_pct", 0)
                health = audit.get("health", "?")
                cs_nums += f" | 映射:{total_m}只(llm {llm_pct}%)[{health}]"
        except Exception:
            cs_nums = ""
    engines.append({
        "name": "🧪 催化筛查", "key": "catalyst_screen",
        "ok": cs_ok, "time": _mtime(cs_path) if cs_ok else "—",
        "nums": cs_nums,
    })

    # catalyst_track
    ct_path = REPORT_DIR / "catalyst" / f"catalyst_track_{today.isoformat()}.md"
    ct_ok = ct_path.exists()
    ct_nums = ""
    if ct_ok:
        try:
            text = ct_path.read_text(encoding="utf-8")
            import re
            m = re.search(r"(\d+)\s*条活性催化.*?走势确认\s*(\d+)\s*条", text)
            if m:
                ct_nums = f"{m.group(1)}活性/{m.group(2)}确认"
            revived = re.findall(r"历史催化复活|历史复活", text)
            if revived:
                ct_nums += " 🔄复活"
            sh = re.search(r"概念预热\s*(\d+)\s*条", text)
            if sh:
                ct_nums += f" 🔥{sh.group(1)}预热"
        except Exception:
            ct_nums = ""
    engines.append({
        "name": "📡 催化跟踪", "key": "catalyst_track",
        "ok": ct_ok, "time": _mtime(ct_path) if ct_ok else "—",
        "nums": ct_nums,
    })

    # primary_synthesis
    ps_path = REPORT_DIR / "feeds" / f"primary_synthesis_{today.isoformat()}.md"
    ps_ok = ps_path.exists()
    ps_nums = ""
    if ps_ok:
        try:
            text = ps_path.read_text(encoding="utf-8")
            import re
            # 共识主题 + 源间分歧 = 板块数
            themes = re.findall(r"^###\s+(?:🔥|📌|⚡)", text, re.MULTILINE)
            ps_nums = f"{len(themes)}板块" if themes else ""
        except Exception:
            ps_nums = ""
    engines.append({
        "name": "🔗 四源交叉", "key": "primary_synthesis",
        "ok": ps_ok, "time": _mtime(ps_path) if ps_ok else "—",
        "nums": ps_nums,
    })

    # wechat_analysis
    wa_path = REPORT_DIR / "wechat_analysis" / f"wechat_analysis_{today.isoformat()}.md"
    wa_ok = wa_path.exists()
    wa_nums = ""
    if wa_ok:
        try:
            text = wa_path.read_text(encoding="utf-8")
            import re
            # 核心主题下 ### 子标题
            core_start = text.find("## 核心主题")
            if core_start >= 0:
                core_section = text[core_start:text.find("\n## ", core_start + 10)]
                themes = re.findall(r"^###\s+", core_section, re.MULTILINE)
                wa_nums = f"{len(themes)}主题" if themes else ""
        except Exception:
            wa_nums = ""
    engines.append({
        "name": "💚 公众号分析", "key": "wechat_analysis",
        "ok": wa_ok, "time": _mtime(wa_path) if wa_ok else "—",
        "nums": wa_nums,
    })

    # zsxq_analysis
    za_path = REPORT_DIR / "zsxq_analysis" / f"zsxq_analysis_{today.isoformat()}.md"
    za_ok = za_path.exists()
    za_nums = ""
    if za_ok:
        try:
            text = za_path.read_text(encoding="utf-8")
            import re
            # 核心主题下 ### 🔥 / ### 📌 子标题
            core_start = text.find("## 核心主题")
            if core_start >= 0:
                core_section = text[core_start:text.find("\n## ", core_start + 10)]
                themes = re.findall(r"^###\s+", core_section, re.MULTILINE)
                za_nums = f"{len(themes)}主题" if themes else ""
        except Exception:
            za_nums = ""
    engines.append({
        "name": "⭐ 星球分析", "key": "zsxq_analysis",
        "ok": za_ok, "time": _mtime(za_path) if za_ok else "—",
        "nums": za_nums,
    })

    # marginal_changes
    mg_path = REPORT_DIR / "marginal" / f"marginal_{today.isoformat()}.md"
    mg_ok = mg_path.exists()
    mg_nums = ""
    if mg_ok:
        try:
            text = mg_path.read_text(encoding="utf-8")
            import re
            m = re.search(r"边际向好\s+\*{0,2}(\d+)\*{0,2}", text)
            up_n = int(m.group(1)) if m else 0
            m2 = re.search(r"边际下滑\s+\*{0,2}(\d+)\*{0,2}", text)
            dn_n = int(m2.group(1)) if m2 else 0
            mg_nums = f"↑{up_n} ↓{dn_n}" if up_n or dn_n else ""
        except Exception:
            mg_nums = ""
    engines.append({
        "name": "📐 边际变化", "key": "marginal",
        "ok": mg_ok, "time": _mtime(mg_path) if mg_ok else "—",
        "nums": mg_nums,
    })

    # industry_deep_read (daily)
    ind_path = REPORT_DIR / "industry" / f"industry_daily_{today.isoformat()}.md"
    ind_ok = ind_path.exists()
    ind_nums = ""
    if ind_ok:
        try:
            text = ind_path.read_text(encoding="utf-8")
            m = re.search(r"(\d+)篇研报\s*\|\s*(\d+)家机构\s*\|\s*(\d+)个行业", text)
            if m:
                ind_nums = f"{m.group(1)}篇/{m.group(2)}机构/{m.group(3)}行业"
        except Exception:
            pass
    engines.append({
        "name": "🏭 行业深研", "key": "industry_deep_read",
        "ok": ind_ok, "time": _mtime(ind_path) if ind_ok else "—",
        "nums": ind_nums,
    })

    # LLM 输出质量审计
    llm_ok = True
    llm_nums = ""
    try:
        from llm_validator import _check_llm_quality
        llm_ok, llm_nums = _check_llm_quality()
    except Exception:
        llm_nums = "校验模块异常"
        llm_ok = False
    engines.append({
        "name": "🤖 LLM输出质量", "key": "llm_quality",
        "ok": llm_ok, "time": "—",
        "nums": llm_nums,
    })

    # === 盘面侧引擎 ===
    srl_ok, srl_time, srl_nums = _sector_rotation_status()
    engines.append({
        "name": "🔄 板块轮动", "key": "sector_rotation",
        "ok": srl_ok, "time": srl_time, "nums": srl_nums,
    })
    mrh_ok, mrh_time, mrh_nums = _market_rhythm_status()
    engines.append({
        "name": "🌊 市场节奏", "key": "market_rhythm",
        "ok": mrh_ok, "time": mrh_time, "nums": mrh_nums,
    })
    sds_ok, sds_time, sds_nums = _similar_days_status()
    engines.append({
        "name": "🔮 相似日匹配", "key": "similar_days",
        "ok": sds_ok, "time": sds_time, "nums": sds_nums,
    })
    lb_ok, lb_time, lb_nums = _leader_backtest_status()
    engines.append({
        "name": "👑 龙头回溯", "key": "leader_backtest",
        "ok": lb_ok, "time": lb_time, "nums": lb_nums,
    })
    lu_ok, lu_time, lu_nums = _limit_up_status()
    engines.append({
        "name": "🎯 涨停分析", "key": "limit_up",
        "ok": lu_ok, "time": lu_time, "nums": lu_nums,
    })

    return engines


def _sector_rotation_log_freshness() -> tuple[bool, str]:
    """检查 sector_rotation_log 表最近数据日期。"""
    try:
        with store._conn() as conn:
            row = conn.execute(
                "SELECT MAX(date) FROM sector_rotation_log WHERE sector != ''"
            ).fetchone()
            if row and row[0]:
                return True, row[0]
    except Exception:
        pass
    return False, "—"


def _sector_rotation_status() -> tuple[bool, str, str]:
    ok, last_date = _sector_rotation_log_freshness()
    if not ok:
        return False, "—", "无数据"
    try:
        freq = sector_frequency(5)
        top3 = "/".join(s["sector"] for s in freq[:3]) if freq else "—"
        return True, last_date, top3
    except Exception:
        return False, last_date, "查询异常"


def _market_rhythm_status() -> tuple[bool, str, str]:
    ok, last_date = _sector_rotation_log_freshness()
    if not ok:
        return False, "—", "无数据"
    try:
        stage = classify_rhythm_stage(last_date)
        stage_cn = {"launch": "+1启动", "relay": "+2接力", "momentum": "主升",
                     "expansion": "扩容", "backflow": "回流", "idle": "空窗"}.get(stage, stage)
        return True, last_date, stage_cn
    except Exception:
        return False, last_date, "查询异常"


def _similar_days_status() -> tuple[bool, str, str]:
    ok, last_date = _sector_rotation_log_freshness()
    if not ok:
        return False, "—", "无数据"
    try:
        result = similar_sectors(last_date, top_n=5)
        pred = result.get("predicted_sectors", [])
        top2 = "/".join(s["sector"] for s in pred[:2]) if pred else "—"
        return True, last_date, top2
    except Exception:
        return False, last_date, "查询异常"


def _leader_backtest_status() -> tuple[bool, str, str]:
    """龙头回溯：聚合历史龙头标的出现频率。"""
    try:
        leaders = leader_frequency(min_appearances=2)
        if not leaders:
            return False, "—", "无数据"
        top3 = "/".join(l["stock"] for l in leaders[:3])
        return True, _sector_rotation_log_freshness()[1], f"{len(leaders)}只(top:{top3})"
    except Exception:
        return False, "—", "查询异常"


def _limit_up_status() -> tuple[bool, str, str]:
    """涨停分析：检查 limit_up_analysis 表今日数据。"""
    try:
        today = date.today().isoformat()
        from store import _conn
        with _conn() as conn:
            row = conn.execute(
                "SELECT COUNT(*) as n, MAX(date) as last_date FROM limit_up_analysis"
            ).fetchone()
            if row and row["n"] > 0:
                t1 = conn.execute(
                    "SELECT COUNT(*) FROM limit_up_analysis WHERE tier='T1' AND date=?",
                    (today,)
                ).fetchone()[0]
                t2 = conn.execute(
                    "SELECT COUNT(*) FROM limit_up_analysis WHERE tier='T2' AND date=?",
                    (today,)
                ).fetchone()[0]
                return True, row["last_date"], f"T1:{t1} T2:{t2}"
            return False, "—", "无数据"
    except Exception:
        return False, "—", "查询异常"


_CONCEPT_STOCKS: dict[str, list[str]] | None = None


def _load_concept_stocks() -> dict[str, list[str]]:
    """加载概念→成分股正向映射（从 multi_concept_map.json 反转，缓存）。"""
    global _CONCEPT_STOCKS
    if _CONCEPT_STOCKS is not None:
        return _CONCEPT_STOCKS
    _CONCEPT_STOCKS = {}
    try:
        import json
        mp = Path(__file__).parent / "data" / "multi_concept_map.json"
        data = json.loads(mp.read_text(encoding="utf-8"))
        for code, concepts in data.get("stocks", {}).items():
            for c in concepts:
                _CONCEPT_STOCKS.setdefault(c, []).append(code)
    except Exception:
        pass
    return _CONCEPT_STOCKS


def _build_code_chain_map() -> dict[str, list[dict]]:
    """构建代码→产业链映射（从 产业链*涨跌幅.xlsx）。"""
    chains = _load_chain_maps()
    code_map: dict[str, list[dict]] = {}
    for plate, l1_map in chains.items():
        for l1, l2_map in l1_map.items():
            for l2, stocks in l2_map.items():
                for s in stocks:
                    code = s["code"]
                    code_map.setdefault(code, []).append({
                        "plate": plate, "l1": l1, "l2": l2,
                    })
    return code_map


def _render_theme_lifecycle(w) -> list[dict]:
    """主题生命周期：从 catalyst_signals 时间轴聚合，追踪主题何时开始、趋势、状态。
    Returns lifecycle rows for downstream advisor use.
    """
    w("## 📅 主题生命周期")
    w()
    try:
        rows = theme_lifecycle(60)
    except Exception as e:
        w(f"_主题生命周期数据不可用: {e}_")
        w()
        return []

    if not rows:
        w("_暂无满足条件的主题_")
        w()
        return []

    state_icon = {"active": "🟢", "confirmed": "✅", "emerging": "🆕", "cooling": "🟡", "dormant": "⚫"}
    w("| 主题 | 状态 | 持续 | 趋势 | 走势 | 首次 | 最近 | 信号 | 均分 |")
    w("|------|:---:|:---:|:---:|:--:|:----:|:----:|:---:|:---:|")
    for r in rows[:20]:
        icon = state_icon.get(r["state"], "·")
        trend_icon = "✅" if r["has_trend"] else "—"
        w(f"| {r['theme']} | {icon} {r['state']} | {r['days_active']}天 | "
          f"{r['trend']} | {trend_icon} | {r['first_date']} | {r['last_date']} | "
          f"{r['signal_count']} | {r['avg_score']:.0f} |")
    w()
    return rows


def _render_theme_advisor(w, lifecycle_rows, chain_rows):
    """规则引擎：交叉生命周期+预期差+个股映射，表格式输出，个股与选中理由逐行对应。"""
    w("### 🧠 主题建议")
    w()
    names = _load_name_cache()
    chain_stocks_all = _load_chain_maps()
    any_output = False

    def _segment_stocks(plate, l1, l2, sort_by: str = "mentions") -> list[dict]:
        l2m = chain_stocks_all.get(plate, {}).get(l1, {})
        key = l2 if l2 and l2 != "-" else list(l2m.keys())[0] if l2m else ""
        stocks = l2m.get(key, []) if key else []
        return sorted(stocks, key=lambda s: -(s.get(sort_by, 0))) if stocks else []

    # ===== 高预期差 =====
    high_gap = [ch for ch in chain_rows
                if ch["verdict"] == "⏳预期差" and ch["expectation_gap"] == "🔴高" and ch["total_mentions"] > 50]
    if high_gap:
        w("#### ⏳ 高预期差（逻辑密集但价格未确认）")
        w()
        w("| 链段 | 个股 | 涨跌 | 提及 | 选中理由 |")
        w("|------|------|:---:|:---:|------|")
        for ch in high_gap[:5]:
            stocks = _segment_stocks(ch["plate"], ch["l1"], ch["l2"], "mentions")
            seg_label = f"{ch['l1']}>{ch['l2']}" if ch['l2'] and ch['l2'] != '-' else ch['l1']
            for i, s in enumerate(stocks[:3]):
                nm = names.get(s["code"], s.get("name", "")) or s.get("name", "?")
                pct = s.get("pct", 0)
                mentions = s.get("mentions", 0)
                if i == 0:
                    reason = f"链段核心标的，全段提及密度{ch['mention_density']:.0f}/只"
                elif pct > 0:
                    reason = "逆势微涨，可能有资金提前布局"
                else:
                    reason = f"链段内提及第{i+1}，随板块回调"
                w(f"| {seg_label} | {nm}({s['code']}) | {pct:+.1f}% | {mentions}次 | {reason} |")
        w()
        any_output = True

    # ===== 三重共振 =====
    resonance = [ch for ch in chain_rows if ch["verdict"] == "🔥共振"]
    if resonance:
        w("#### 🔥 三重共振（逻辑+价格+走势同步确认）")
        w()
        w("| 链段 | 走势信号 | 个股 | 涨跌 | 操作建议 |")
        w("|------|:-------:|------|:---:|------|")
        for ch in resonance[:3]:
            stocks = _segment_stocks(ch["plate"], ch["l1"], ch["l2"], "mentions")
            seg_label = f"{ch['l1']}>{ch['l2']}" if ch['l2'] and ch['l2'] != '-' else ch['l1']
            for i, s in enumerate(stocks[:2]):
                nm = names.get(s["code"], s.get("name", "")) or s.get("name", "?")
                advice = "已共振，不建议追高，等回调至均线" if i == 0 else "同上"
                w(f"| {seg_label} | {ch['trend_signal']} | {nm}({s['code']}) | {s.get('pct',0):+.1f}% | {advice} |")
        w()
        any_output = True

    # ===== 走势先行 =====
    trend_lead = [ch for ch in chain_rows if ch["verdict"] == "👀走势先行"]
    if trend_lead:
        w("#### 👀 走势先行（价格已动但逻辑稀疏，需深挖）")
        w()
        w("| 链段 | 个股 | 涨跌 | 逻辑提及 | 待验证 |")
        w("|------|------|:---:|:-------:|------|")
        for ch in trend_lead[:4]:
            stocks = _segment_stocks(ch["plate"], ch["l1"], ch["l2"], "pct")
            seg_label = f"{ch['l1']}>{ch['l2']}" if ch['l2'] and ch['l2'] != '-' else ch['l1']
            for i, s in enumerate(stocks[:2]):
                nm = names.get(s["code"], s.get("name", "")) or s.get("name", "?")
                verify = "翻最近5日公告+研报，确认催化来源" if i == 0 else "同上"
                w(f"| {seg_label} | {nm}({s['code']}) | {s.get('pct',0):+.1f}% | {ch['total_mentions']}次 | {verify} |")
        w()
        any_output = True

    # ===== 主升确认 =====
    confirmed = [lr for lr in lifecycle_rows
                 if lr["state"] == "confirmed" and lr["trend"] == "🔥加剧"]
    if confirmed:
        w("#### 📈 主升确认（走势已确认+逻辑仍在加剧）")
        w()
        w("| 主题 | 链段 | 个股 | 涨跌 | 状态 |")
        w("|------|------|------|:---:|------|")
        for lr in confirmed[:3]:
            related = [ch for ch in chain_rows
                      if ch["plate"] == lr["theme"] or lr["theme"] in ch["plate"]]
            shown = set()
            for ch in related[:3]:
                seg_label = f"{ch['l1']}>{ch['l2']}" if ch['l2'] and ch['l2'] != '-' else ch['l1']
                for s in _segment_stocks(ch["plate"], ch["l1"], ch["l2"], "mentions")[:2]:
                    if s["code"] in shown:
                        continue
                    shown.add(s["code"])
                    nm = names.get(s["code"], s.get("name", "")) or s.get("name", "?")
                    status = ch["verdict"] if ch["verdict"] != "—" else "跟踪中"
                    w(f"| {lr['theme']} | {seg_label} | {nm}({s['code']}) | {s.get('pct',0):+.1f}% | {status} |")
        w()
        any_output = True

    if not any_output:
        w("_当前无显著交叉信号_")
    w()


def _render_dual_track(w):
    """双轨选股：FEV 格雷厄姆轨 + G-Factor 费雪轨 分席输出。"""
    w("## 🏆 双轨选股（FEV + G-Factor）")
    w()
    try:
        result = dual_track_screen(fev_top_n=7, g_per_dim=3, min_fev=15, min_g=6)
    except Exception as e:
        w(f"_双轨数据不可用: {e}_")
        w()
        return

    ta = result["track_a"]
    tb = result["track_b"]

    # 轨A: FEV
    w(f"### 轨A: {ta['label']}（{ta['count']}只）")
    w()
    if ta["stocks"]:
        w("| 代码 | 名称 | F | E | V | FEV | 备注 |")
        w("|------|------|:--:|:--:|:--:|:---:|------|")
        for s in ta["stocks"]:
            w(f"| {s['code']} | {s['name']} | {s['f_score']} | {s['e_score']} | "
              f"{s['v_score']} | **{s['fev_total']}** | {s.get('f_note','')[:30]} |")
    else:
        w("_无达标标的_")
    w()

    # 轨B: G-Factor（四维按标的去重合并展示）
    w(f"### 轨B: {tb['label']}（{tb['count']}只）")
    w()
    code_rows: dict[str, dict] = {}
    for dim_name, entries in tb["dims"].items():
        dim_short = dim_name.split("_")[1] if "_" in dim_name else dim_name
        for e in entries:
            c = e["code"]
            if c not in code_rows:
                code_rows[c] = {"name": e["name"], "scores": {}}
            code_rows[c]["scores"][dim_short] = e["score"]
    if code_rows:
        w("| 代码 | 名称 | 成长质量 | 催化密度 | 叙事强度 | 机构动量 |")
        w("|------|------|:------:|:------:|:------:|:------:|")
        for code, info in code_rows.items():
            sc = info["scores"]
            w(f"| {code} | {info['name']} | {sc.get('成长质量','—')} | "
              f"{sc.get('催化密度','—')} | {sc.get('叙事强度','—')} | {sc.get('机构动量','—')} |")
    else:
        w("_无达标标的_")
    w()

    # 双轨交集
    if result["overlap"]:
        overlap_codes = result["overlap"]
        w(f"### 🎯 双轨交集（{len(overlap_codes)}只）")
        w()
        names = _load_name_cache()
        named = [f"{names.get(c, '')}({c})" for c in overlap_codes]
        w(" · ".join(named))
        w()


def _render_price_signals(w, last_trade: str):
    """盘面侧信号：概念热度（commonality_cache 自动化数据） + 持续性/相似日（sector_rotation_log 历史数据）。"""
    w("## 📈 盘面概念信号（自下而上）")
    w()
    names = _load_name_cache()

    # 1. 概念热度 — 基于 commonality_cache（每日自动产出）
    try:
        from catalyst_tracker import _check_concept_heat, _load_concept_baseline, _load_today_concept_counts
        concept_heat = _check_concept_heat(last_trade)
        baseline = _load_concept_baseline()
        concept_stocks = _load_concept_stocks()

        if concept_heat:
            counts = _load_today_concept_counts(last_trade)
            w("### 🔥 今日概念热度（成分股协同异动 Top 12）")
            w()
            w("| 概念 | 热度比 | 强势/总量 | 代表标的 |")
            w("|------|:-----:|:-----:|------|")
            for concept, ratio in sorted(concept_heat.items(), key=lambda x: -x[1])[:12]:
                base = baseline.get(concept, 0)
                cnt = counts.get(concept, 0)
                stocks = concept_stocks.get(concept, [])[:4]
                named = []
                for s in stocks:
                    nm = names.get(s, "")
                    named.append(f"{nm}({s})" if nm else s)
                stocks_str = " ".join(named) if named else "—"
                w(f"| {concept} | {ratio*100:.0f}% | {cnt}/{base} | {stocks_str} |")
            w()
        else:
            w(f"_今日无概念触发热度阈值_")
            w()
    except Exception as e:
        w(f"_概念热度: {e}_")
        w()

    # 2. 持续性板块（sector_rotation_log 历史数据）+ 核心个股
    try:
        pers = sector_persistence(5)
        if pers:
            w("### 📊 历史持续性板块（最长连续出现）")
            w()
            w("| 板块 | 最长连续 | 平均连续 | 轮次 | 核心个股 |")
            w("|------|:-----:|:-----:|:---:|------|")
            for s in pers[:10]:
                top = sector_stocks(s["sector"], 5)
                named = []
                for st in top:
                    nm = names.get(st["stock"], "")
                    named.append(f"{nm}({st['stock']})" if nm else st["stock"])
                stocks_str = " ".join(named[:3]) if named else "—"
                w(f"| {s['sector']} | {s['max_streak']} | {s['avg_streak']} | {s['runs']} | {stocks_str} |")
            w()
    except Exception:
        pass

    # 3. 相似日预测（sector_rotation_log 历史数据）
    try:
        result = similar_sectors(last_trade, top_n=5, lookahead=3)
        pred = result.get("predicted_sectors", [])
        if pred:
            w("### 🔮 相似日预测板块")
            w()
            w("| 板块 | 加权得分 |")
            w("|------|:-----:|")
            for s in pred[:8]:
                w(f"| {s['sector']} | {s['score']} |")
            current_sectors = set(result.get("target_sectors", []))
            new_sectors = [s for s in pred if s["sector"] not in current_sectors]
            if new_sectors:
                w()
                w("*新方向提示*: " + ", ".join(s["sector"] for s in new_sectors[:5]))
            w()
    except Exception:
        pass


def _source_funnel(source: str, msg: str) -> tuple[str, str, str]:
    """从 collector 消息中提取吞吐量三阶段: (采集, 初筛, 深度产出)。"""
    import re
    today = date.today().isoformat()

    # 深度分析管线 — 三阶段拆解
    if source == "announcement_deep_read":
        collect = "—"
        filtered = ""
        deep = ""
        try:
            with store._conn() as conn:
                total = conn.execute(
                    "SELECT COUNT(*) FROM deep_read_results WHERE date=?",
                    (today,),
                ).fetchone()[0]
                a60 = conn.execute(
                    "SELECT COUNT(*) FROM deep_read_results WHERE date=? AND total_score>=60",
                    (today,),
                ).fetchone()[0]
            if total:
                deep = f"{total}篇LLM（{a60}≥60分）"
        except Exception:
            pass
        return (collect, filtered or "—", deep or "—")

    if source == "research_deep_read":
        d = REPORT_DIR / "research_dossiers"
        try:
            n = len(list(d.glob("*.md"))) if d.exists() else 0
            return ("—", "信号检测", f"{n}份档案")
        except Exception:
            pass
        return ("—", "—", "—")

    if source == "sentiment_track":
        m = re.search(r"(\d+)存档", msg)
        if m:
            return ("—", "调研+互动+业绩", f"{m.group(1)}只存档")
        return ("—", "—", "—")

    if source == "industry_deep_read":
        ind_path = REPORT_DIR / "industry" / f"industry_daily_{today}.md"
        if ind_path.exists():
            try:
                text = ind_path.read_text(encoding="utf-8")
                m = re.search(r"(\d+)篇研报\s*\|\s*(\d+)家机构\s*\|\s*(\d+)个行业", text)
                if m:
                    return (f"{m.group(1)}篇", f"{m.group(2)}家机构", f"{m.group(3)}行业研判")
            except Exception:
                pass
        return ("—", "—", "—")

    if source == "news_signals":
        sig_path = REPORT_DIR / "feeds" / f"news_signals_{today}.md"
        if sig_path.exists():
            try:
                text = sig_path.read_text(encoding="utf-8")
                m = re.search(r"(\d+)条边际信号", text)
                if m:
                    return ("新闻列表", "Haiku扫描", f"{m.group(1)}条信号")
            except Exception:
                pass
        return ("—", "—", "—")

    if source == "catalyst_tracker":
        ct_path = REPORT_DIR / "catalyst" / f"catalyst_track_{today}.md"
        if ct_path.exists():
            try:
                text = ct_path.read_text(encoding="utf-8")
                m = re.search(r"(\d+)\s*条活性催化.*?确认\s*(\d+)\s*条", text)
                if m:
                    return (f"{m.group(1)}活性", "Redis扫描", f"{m.group(2)}确认")
            except Exception:
                pass
        return ("—", "—", "—")

    # 普通采集源 — 提取关键数字
    num_str = ""
    for pat, unit in [
        (r"(\d+)\s*篇", "篇"),
        (r"命中(\d+)", "条命中"),
        (r"(\d+)\s*只\s*成功", "只"),
        (r"(\d+)\s*条", "条"),
        (r"新增\s*(\d+)", "条↑"),
        (r"成功(\d+)", "只"),
    ]:
        m = re.search(pat, msg)
        if m:
            n = int(m.group(1))
            if "无新" in msg or "无帖子" in msg:
                num_str = f"{n}{unit}"
            else:
                num_str = f"{n}{unit}"
            break

    if not num_str:
        if "超时" in msg:
            num_str = "⏰ 超时"
        else:
            num_str = msg[:25] if msg else "—"

    return (num_str, "—", "—")


def _render_output_checklist(w, today_str: str = ""):
    """渲染「📋 输出清单」— 所有预期日产出的一眼状态。"""
    from output_audit import check_all

    results = check_all()
    now = datetime.now()

    w("## 📋 输出清单")
    w("")
    w("| 产出 | 状态 | 产出时间 | 目标时间 | 落后 | 管线 |")
    w("|------|:----:|:------:|:------:|:----:|------|")

    ok_count = 0
    stale_count = 0
    for r in results:
        if r["ok"] and not r.get("stale"):
            icon = "✅"
            ok_count += 1
        elif r.get("stale"):
            icon = "⚠️"
            stale_count += 1
        else:
            icon = "❌"

        mtime_str = r["mtime_str"]
        target_dt_str = f"{r['target_date'][5:]} {r['deadline']}"  # MM-DD HH:MM
        db_flag = " 💾" if r["is_db"] else ""

        delay_str = ""
        try:
            dl = datetime.strptime(f"{r['target_date']} {r['deadline']}", "%Y-%m-%d %H:%M")
            gap_h = (now - dl).total_seconds() / 3600
            if not r["ok"] and gap_h > 0:
                delay_str = f"{gap_h:.0f}h"
                icon = "🔴"
            elif not r["ok"] and gap_h > -0.5:
                delay_str = "⏳"
            elif r.get("stale") and gap_h > 0:
                delay_str = f"{gap_h:.0f}h"
        except ValueError:
            pass

        stale_note = " 📂过期" if r.get("stale") else ""
        w(f"| {r['name']}{db_flag} | {icon} | {mtime_str} | {target_dt_str} | {delay_str} | {r['pipeline']}{stale_note} |")

    w("")
    missing = len(results) - ok_count - stale_count
    if missing == 0 and stale_count == 0:
        w(f"✅ **全部 {len(results)} 项产出就绪**")
    else:
        parts = []
        if ok_count:
            parts.append(f"{ok_count} 正常")
        if stale_count:
            parts.append(f"{stale_count} ⚠️过期")
        if missing:
            missing_names = "、".join(r["name"] for r in results if not r["ok"])
            parts.append(f"{missing} ❌缺失: {missing_names}")
        w("⚠️ **" + "，".join(parts) + "**")
    w("")


def _load_name_cache() -> dict[str, str]:
    """加载全市场股票名称缓存。"""
    name_map = {}
    try:
        import json
        cache_path = Path(__file__).parent / "data" / "stock_codes.json"
        cache = json.loads(cache_path.read_text(encoding="utf-8"))
        for item in cache.get("codes", []):
            name_map[item["code"]] = item["name"]
    except Exception:
        pass
    return name_map


def _recent_files(directory: Path, prefix: str, days: int = 7) -> list[Path]:
    """Return files in <directory> starting with <prefix>, filtered to recent <days> dates."""
    if not directory.exists():
        return []
    recent = {(date.today() - timedelta(days=d)).strftime("%Y-%m-%d") for d in range(days)}
    result = []
    for fp in sorted(directory.glob(f"{prefix}*.md")):
        m = re.search(r"(\d{4}-\d{2}-\d{2})", fp.name)
        if m and m.group(1) in recent:
            result.append(fp)
    return result


def _aggregate_signals(dr: dict) -> dict:
    """6源信号矩阵：公告deep_read + 研报 + 新闻信号 + 行业深研 + 社交源 + 催化信号。

    权重原则：
      - 研报是机构专业研究，1篇即进初步筛选，2篇+其他源可进深度分析
      - 社交源（公众号/微博）是人工精选信息源，提及即进初步筛选
      - 行业深研提到代表跨行业共识，权重高
      - 公告deep_read通过Stage1硬筛选的至少值10分（已是精选）
      - 新闻边际信号经Haiku扫描已过滤无关内容
    """
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    names = _load_name_cache()
    scores: dict[str, dict] = {}

    def _ensure(code):
        code = str(code).zfill(6)
        if code not in scores:
            scores[code] = {"code": code, "name": names.get(code, ""),
                            "dr_score": 0, "dr_count": 0,
                            "rpt_count": 0, "ns_count": 0,
                            "ind_count": 0, "social_count": 0,
                            "cat_score": 0, "cat_count": 0,
                            "sources": 0, "total": 0}

    # ====== 1. 公告 deep_read（>=20 分即纳入，<20 也给基础分） ======
    try:
        with store._conn() as conn:
            rows = conn.execute(
                "SELECT code, MAX(total_score) as ms, COUNT(*) as cnt "
                "FROM deep_read_results WHERE date >= ? AND total_score >= 10 "
                "GROUP BY code",
                (week_ago,),
            ).fetchall()
        for r in rows:
            _ensure(r["code"])
            s = scores[r["code"]]
            s["dr_score"] = max(r["ms"], 10)  # 至少值10分（通过了Stage1硬筛选）
            s["dr_count"] = r["cnt"]
            s["sources"] += 1
    except Exception:
        pass

    # ====== 2. 研报采集（每篇研报都算信号） ======
    try:
        with store._conn() as conn:
            rows = conn.execute(
                "SELECT code, COUNT(*) as cnt FROM research_reports "
                "WHERE report_date >= ? GROUP BY code",
                (week_ago,),
            ).fetchall()
        for r2 in rows:
            code = str(r2["code"]).zfill(6)
            if len(code) != 6:
                continue
            _ensure(code)
            s = scores[code]
            s["rpt_count"] = r2["cnt"]
            s["sources"] += 1
    except Exception:
        pass

    # ====== 3. 新闻边际信号（Haiku提取的每条边际变化） ======
    for d_dir in range(7):
        d_str = (date.today() - timedelta(days=d_dir)).isoformat()
        fp = REPORT_DIR / "feeds" / "news_signals" / f"news_signals_{d_str}.md"
        if not fp.exists():
            continue
        try:
            for line in fp.read_text(encoding="utf-8").split("\n"):
                if line.startswith("| ") and line[2:3].isdigit():
                    parts = [p.strip() for p in line.split("|")]
                    if len(parts) >= 3:
                        code = parts[1]
                        if len(code) == 6 and code.isdigit():
                            _ensure(code)
                            scores[code]["ns_count"] += 1
                            scores[code]["sources"] += 1
        except Exception:
            pass

    # ====== 4. 行业深研 — 从 industry_daily 报告中提取提及标的 ======
    for fp in _recent_files(REPORT_DIR / "industry", "industry_daily_"):
        try:
            text = fp.read_text(encoding="utf-8")
            for code in extract_codes_from_text(text):
                _ensure(code)
                scores[code]["ind_count"] += 1
                scores[code]["sources"] += 1
        except Exception:
            pass

    # ====== 5. 社交源：公众号分析 + 微博 ======
    for fp in _recent_files(REPORT_DIR / "wechat_analysis", "wechat_analysis_"):
        try:
            text = fp.read_text(encoding="utf-8")
            for code in extract_codes_from_text(text):
                _ensure(code)
                scores[code]["social_count"] += 1
                scores[code]["sources"] += 1
        except Exception:
            pass
    for fp in _recent_files(REPORT_DIR / "feeds" / "weibo", "weibo_"):
        try:
            text = fp.read_text(encoding="utf-8")
            for code in extract_codes_from_text(text):
                _ensure(code)
                scores[code]["social_count"] += 1
                scores[code]["sources"] += 1
        except Exception:
            pass

    # ====== 6. 催化信号（星球+deep_read产出） ======
    try:
        with store._conn() as conn:
            rows = conn.execute(
                "SELECT mentioned_codes, MAX(actionability) as ms, COUNT(*) as cnt "
                "FROM catalyst_signals WHERE date >= ? AND actionability >= 10 "
                "GROUP BY mentioned_codes",
                (week_ago,),
            ).fetchall()
        for r3 in rows:
            code = str(r3["mentioned_codes"] or "").strip()
            if not code or len(code) != 6 or not code.isdigit():
                continue
            _ensure(code)
            s = scores[code]
            s["cat_score"] = max(s["cat_score"], r3["ms"])
            s["cat_count"] = r3["cnt"]
            s["sources"] += 1
    except Exception:
        pass

    # ====== 计算复合分 + 交叉源加分 ======
    for s in scores.values():
        # 重新计算来源数（基于字段是否有数据，而非逐条累加）
        src_count = 0
        if s["dr_count"] > 0: src_count += 1
        if s["rpt_count"] > 0: src_count += 1
        if s["ns_count"] > 0: src_count += 1
        if s["ind_count"] > 0: src_count += 1
        if s["social_count"] > 0: src_count += 1
        if s["cat_count"] > 0: src_count += 1
        s["sources"] = src_count

        base = (s["dr_score"] * 1.0           # 公告 deep_read: max(score,10) → 10~85
                + s["rpt_count"] * 20          # 研报: 20分/篇，1篇即进初步筛选
                + s["ns_count"] * 8            # 新闻边际信号: 8分/条，Haiku已初筛
                + s["ind_count"] * 10          # 行业深研提及: 10分/次
                + s["social_count"] * 10       # 社交源提及: 10分/次（公众号/微博）
                + s["cat_score"] * 0.8)        # 催化行动分
        cross = 0
        if s["sources"] >= 6: cross = 65
        elif s["sources"] >= 5: cross = 50
        elif s["sources"] >= 4: cross = 35
        elif s["sources"] >= 3: cross = 20
        elif s["sources"] >= 2: cross = 10
        s["total"] = int(base + cross)
        if not s["name"]:
            s["name"] = names.get(s["code"], "")

    ranked = sorted(scores.values(), key=lambda x: -x["total"])

    deep = [s for s in ranked if s["total"] >= 80]    # 深度分析: ≥80
    watch = [s for s in ranked if 20 <= s["total"] < 80]  # 初步关注: ≥20（1篇研报或2次社交提及即达标）

    return {"deep": deep[:20], "watch": watch[:20]}


def _hot_themes() -> list[dict]:
    """多源交叉主题热度：catalyst_signals + 行业研报 + 星球 + 公众号。"""
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    themes: dict[str, dict] = {}
    names = _load_name_cache()

    def _add_theme(keyword, source, stocks=None):
        if not keyword or keyword in ("—", "other", "deep_read", "tech_breakthrough", "policy_change"):
            return
        if keyword not in themes:
            themes[keyword] = {"theme": keyword, "count": 0, "sources": set(), "stocks": set()}
        themes[keyword]["count"] += 1
        themes[keyword]["sources"].add(source)
        if stocks:
            for s in stocks:
                if re.match(r"\d{6}", str(s)):
                    themes[keyword]["stocks"].add(str(s))

    KW_MAP = [
        ("PCB",), ("MLCC",), ("AI芯片",), ("GPU",), ("HBM",), ("算力",), ("存储",), ("液冷",),
        ("光模块",), ("CPO",), ("先进封装",), ("半导体设备",), ("碳化硅",), ("覆铜板",),
        ("断供",), ("缺货",), ("供不应求",), ("涨价",), ("产能扩张",),
        ("新能源",), ("锂电",), ("光伏",), ("储能",), ("机器人",), ("军工",), ("航天",),
        ("邮轮", "VLCC", "航运"), ("稀土",), ("收购",), ("重组",), ("跨界",),
        ("技术突破",), ("业绩暴",), ("股权激励",), ("减持",),
        ("华为",), ("特斯拉",), ("比亚迪",), ("宁德",),
    ]

    # 1. catalyst_signals
    try:
        with store._conn() as conn:
            rows = conn.execute(
                "SELECT catalyst_type, thesis, mentioned_codes FROM catalyst_signals "
                "WHERE date >= ? AND actionability >= 20",
                (week_ago,),
            ).fetchall()
    except Exception:
        rows = []
    for r in rows:
        thesis = (r["thesis"] or "") + " " + (r["catalyst_type"] or "")
        found = False
        for kw_group in KW_MAP:
            for kw in kw_group:
                if kw in thesis:
                    codes = [c.strip() for c in str(r["mentioned_codes"] or "").split(",") if c.strip()]
                    _add_theme(kw_group[0], "催化", codes)
                    found = True
                    break
            if found:
                break
        if not found and "减持" not in thesis and "订单" not in thesis:
            _add_theme(r["catalyst_type"], "催化")

    # 2. 行业深研合成报告
    for fp in _recent_files(REPORT_DIR / "industry", "industry_daily_"):
        try:
            text = fp.read_text(encoding="utf-8")
            in_consensus = False
            for line in text.split("\n"):
                if "今日共识方向" in line:
                    in_consensus = True
                    continue
                if in_consensus and line.startswith("##"):
                    break
                if in_consensus and line.startswith("- "):
                    direction = line[2:].strip()
                    for kw_group in KW_MAP:
                        for kw in kw_group:
                            if kw in direction:
                                codes = list(extract_codes_from_text(direction))
                                _add_theme(kw_group[0], "行业深研", codes)
                                break
        except Exception:
            pass

    # 3. 公众号分析
    for fp in _recent_files(REPORT_DIR / "wechat_analysis", "wechat_analysis_"):
        try:
            text = fp.read_text(encoding="utf-8")
            for m in re.finditer(r"###\s+(.+?)(?:\n|$)", text):
                title = m.group(1)
                for kw_group in KW_MAP:
                    for kw in kw_group:
                        if kw in title:
                            codes = list(extract_codes_from_text(text[m.start():m.start()+500]))
                            _add_theme(kw_group[0], "公众号", codes)
                            break
        except Exception:
            pass

    # 4. 星球分析
    for fp in _recent_files(REPORT_DIR / "zsxq_analysis", "zsxq_analysis_"):
        try:
            text = fp.read_text(encoding="utf-8")
            for m in re.finditer(r"###\s+(.+?)(?:\n|$)", text):
                title = m.group(1)
                for kw_group in KW_MAP:
                    for kw in kw_group:
                        if kw in title:
                            codes = list(extract_codes_from_text(text[m.start():m.start()+500]))
                            _add_theme(kw_group[0], "星球", codes)
                            break
        except Exception:
            pass

    result = []
    # Cross-reference with chain maps for sub-sector tags
    chains = _load_chain_maps()
    tag_map: dict[str, list[str]] = {}
    if chains:
        for plate, l1_map in chains.items():
            for l1, l2_map in l1_map.items():
                for l2, stocks in l2_map.items():
                    if not stocks:
                        continue
                    seg_text = f"{plate} {l1} {l2} " + " ".join(
                        f"{s['name']} {s['reason']}" for s in stocks[:3]
                    )
                    for kw_group in KW_MAP:
                        for kw in kw_group:
                            if kw in seg_text:
                                tag = f"{l1}>{l2}" if l2 and l2 != '-' else l1
                                plate_tag = f"{plate}>{tag}"
                                if kw_group[0] not in tag_map:
                                    tag_map[kw_group[0]] = []
                                if plate_tag not in tag_map[kw_group[0]]:
                                    tag_map[kw_group[0]].append(plate_tag)
                                break

    for t in themes.values():
        stock_list = list(t["stocks"])
        named = []
        for s in stock_list[:3]:
            nm = names.get(s, "")
            named.append(f"{nm}({s})" if nm else s)
        t["stocks"] = " ".join(named)
        t["heat"] = min(t["count"] // 5 + 1, 3)
        t["source_list"] = "+".join(sorted(t["sources"]))
        t["chain_tags"] = " / ".join(tag_map.get(t["theme"], [])[:3])
        result.append(t)

    result.sort(key=lambda x: -x["count"])
    return result[:15]


def _discover_chain_xlsx() -> list[Path]:
    """Auto-discover chain map XLSX files: *产业链*涨跌幅.xlsx"""
    project_root = Path(__file__).parent.parent
    xlsx_files = []
    for pattern in ["*产业链*涨跌幅.xlsx", "*产业链*.xlsx"]:
        for fp in project_root.glob(pattern):
            if fp not in xlsx_files:
                xlsx_files.append(fp)
    return sorted(xlsx_files)


def _load_chain_maps() -> dict:
    """Load chain maps from XLSX files into {板块: {层级1: {层级2: [(code, name, pct, mcap, mentions, reason), ...]}}}"""
    try:
        import openpyxl
    except ImportError:
        return {}

    chains: dict[str, dict] = {}
    name_cache = _load_name_cache()

    for fp in _discover_chain_xlsx():
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception:
            continue
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # forward-fill merged cells
            cur_plate = cur_l1 = cur_l2 = ""
            for row in rows:
                if not row or len(row) < 11:
                    continue
                # Forward-fill plate/l1/l2 (XLSX has merged cells)
                plate = str(row[0] or "").strip() or cur_plate
                l1 = str(row[1] or "").strip() or cur_l1
                l2 = str(row[2] or "").strip() or cur_l2 or "-"
                cur_plate, cur_l1, cur_l2 = plate, l1, l2

                code_raw = str(row[3] or "").strip()
                name = str(row[4] or "").strip()
                # Parse percentage: "-1.76859%" or "-4.20%" or "2.22%"
                pct_str = str(row[6] or "0").replace("%", "").strip()
                try:
                    pct = float(pct_str)
                except (ValueError, TypeError):
                    pct = 0.0
                try:
                    mcap = float(row[7] or 0)
                except (ValueError, TypeError):
                    mcap = 0.0
                try:
                    mentions = int(row[8] or 0)
                except (ValueError, TypeError):
                    mentions = 0
                reason = str(row[10] or "").strip()

                # Normalize code: 300398.SZ -> 300398
                code = code_raw.replace(".SZ", "").replace(".SH", "").replace(".BJ", "")
                if not code.isdigit() or len(code) != 6:
                    continue
                if not name:
                    name = name_cache.get(code, "")

                if plate not in chains:
                    chains[plate] = {}
                if l1 not in chains[plate]:
                    chains[plate][l1] = {}
                if l2 not in chains[plate][l1]:
                    chains[plate][l1][l2] = []

                chains[plate][l1][l2].append({
                    "code": code, "name": name, "pct": pct,
                    "mcap": mcap, "mentions": mentions, "reason": reason,
                })

    return chains


def _compute_trend_signal(codes: list[str], concept_heat: dict[str, float],
                          concept_stocks: dict[str, list[str]]) -> str:
    """链环节标的与今日热门概念的交集 → 走势信号。"""
    if not concept_heat or not codes:
        return "—"
    best_concept, best_overlap = None, 0
    for concept, ratio in concept_heat.items():
        c_stocks = set(concept_stocks.get(concept, []))
        overlap = sum(1 for c in codes if c in c_stocks)
        if overlap > best_overlap:
            best_overlap, best_concept = overlap, concept
    if best_overlap >= 2 and best_concept:
        return f"🔥{best_concept}({concept_heat[best_concept]*100:.0f}%)"
    return "—"


def _compute_expectation_gap(mention_density: float, avg_pct: float) -> str:
    """逻辑提及密度 vs 价格反应 → 预期差。
    提及多但涨得少=逻辑未被充分定价；提及多且涨得多=逻辑已消化。"""
    if mention_density < 1:
        return "—"
    gap_ratio = mention_density / max(abs(avg_pct), 0.5)
    if gap_ratio > 15:
        return "🔴高"
    elif gap_ratio > 5:
        return "🟡中"
    return "🟢低"


def _compute_verdict(mention_density: float, avg_pct: float, trend_signal: str) -> str:
    """双轨交叉综合判断：
    🔥共振=逻辑密集+价格上涨+走势验证（三重确认）
    ⏳预期差=逻辑密集+价格未动（未被充分定价，机会窗口）
    👀走势先行=逻辑稀疏+价格上涨+走势先行（需深挖逻辑支撑）
    """
    has_logic = mention_density >= 3
    has_price = avg_pct > 2
    has_trend = trend_signal != "—"
    if has_logic and has_price and has_trend:
        return "🔥共振"
    if has_logic and not has_price:
        return "⏳预期差"
    if not has_logic and has_price and has_trend:
        return "👀走势先行"
    return "—"


def _chain_heat() -> list[dict]:
    """Calculate heat per chain segment: 板块 → 层级1 → 层级2。
    叠加概念走势信号 + 预期差 + 双轨综合判断。
    """
    chains = _load_chain_maps()
    if not chains:
        return []

    concept_heat = {}
    try:
        from catalyst_tracker import _check_concept_heat
        last_trade = _last_trading_day().isoformat()
        concept_heat = _check_concept_heat(last_trade)
    except Exception:
        pass
    concept_stocks = _load_concept_stocks()

    result = []
    for plate, l1_map in chains.items():
        for l1, l2_map in l1_map.items():
            for l2, stocks in l2_map.items():
                if not stocks:
                    continue
                count = len(stocks)
                codes = [s["code"] for s in stocks]
                avg_pct = sum(s["pct"] for s in stocks) / count
                total_mentions = sum(s["mentions"] for s in stocks)
                best = max(stocks, key=lambda s: (s["mentions"], s["pct"]))
                mention_density = total_mentions / count if count > 0 else 0
                momentum = "🔥" if avg_pct > 3 else ("🟢" if avg_pct > 0 else "🔴")

                trend = _compute_trend_signal(codes, concept_heat, concept_stocks)
                gap = _compute_expectation_gap(mention_density, avg_pct)
                verdict = _compute_verdict(mention_density, avg_pct, trend)

                result.append({
                    "plate": plate, "l1": l1, "l2": l2,
                    "count": count, "avg_pct": round(avg_pct, 1),
                    "total_mentions": total_mentions,
                    "mention_density": round(mention_density, 1),
                    "top_code": best["code"], "top_name": best["name"],
                    "top_mentions": best["mentions"], "top_pct": best["pct"],
                    "top_reason": best["reason"][:80] if best["reason"] else "",
                    "momentum": momentum,
                    "trend_signal": trend,
                    "expectation_gap": gap,
                    "verdict": verdict,
                })

    verdict_order = {"🔥共振": 0, "⏳预期差": 1, "👀走势先行": 2, "—": 3}
    result.sort(key=lambda x: (
        verdict_order.get(x["verdict"], 4),
        -(abs(x["avg_pct"]) * max(x["mention_density"], 1) * x["count"])
    ))
    return result


def generate(today_str: str = "") -> str:
    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    today = today_str or date.today().isoformat()
    last_trade = _last_trading_day().isoformat()
    L = []
    def w(s=""): L.append(s)

    w(f"# 📊 每日仪表盘 {today}")
    w()
    w(f"> 生成时间: {now_ts} | [复盘](daily_review/reports/review/review_{last_trade}.md) | [建议](daily_review/reports/advice/advice_{today}.md) | [深研档案](daily_review/reports/deep_read/) | [个股档案](daily_review/reports/research_dossiers/)")
    w()

    # === 五维状态 ===
    w("## 五维信息源")
    w()
    w("| 维度 | 采集 | 数据 | 信号 | 成本 |")
    w("|------|:---:|------|-----:|:----:|")
    collectors = {r["source"]: r for r in _collector_status()}
    dr = _deep_read_weekly()
    dc = _dossier_count()
    _generate_dossier_index()

    for dim in DIMENSIONS:
        cs = collectors.get(dim["collector"], {})
        icon = _status_icon(cs.get("status", "unknown"))
        freshness, _ = _data_freshness(dim["data_source"])

        # 信号
        if dim["signal_query"] == "announce_signal":
            sig = f">=60: {dr['total_a60']}条"
        elif dim["signal_query"] == "dossier":
            sig = f"{dc}份档案"
        else:
            sig = "—"

        w(f"| {dim['icon']} {dim['name']} | {icon} | {freshness} | {sig} | {dim['cost']} |")

    w()

    # === 采集管线 ===
    w("## 采集管线")
    w()
    w("| 源 | 状态 | 数据日 | 时间 | 采集 | 初筛 | 深度产出 |")
    w("|----|:----:|:-----:|:---:|------|------|---------|")
    for cs in _collector_status():
        icon = _status_icon(cs.get("status", "unknown"))
        label = SOURCE_LABELS.get(cs["source"], cs["source"])
        last = cs.get("last_date", "") or "—"
        rtime = cs.get("run_time", "") or "—"
        c1, c2, c3 = _source_funnel(cs["source"], cs.get("message", ""))
        w(f"| {label} | {icon} | {last} | {rtime} | {c1} | {c2} | {c3} |")
    w()

    # === 分析引擎 ===
    w("## 分析引擎")
    w()
    w("| 引擎 | 状态 | 产出时间 | 关键数字 |")
    w("|------|:----:|:------:|---------|")
    for eng in _engine_status():
        icon = "✅" if eng["ok"] else "❌"
        w(f"| {eng['name']} | {icon} | {eng['time']} | {eng['nums'] or '—'} |")
    w()

    # === 输出清单 ===
    _render_output_checklist(w, today)

    # === 异常警报（可操作） ===
    alerts = []
    for cs in _collector_status():
        if cs.get("status") in ("error", "timeout", "unknown"):
            label = SOURCE_LABELS.get(cs["source"], cs["source"])
            msg = cs.get("message", "无详情")
            actions = ALERT_ACTIONS.get(cs["source"], {})
            # 匹配排查指引
            hint = ""
            for kw, action in actions.items():
                if kw in msg:
                    hint = action
                    break
            if not hint and cs["status"] == "timeout":
                hint = "调大 daily_collect.COLLECTOR_TIMEOUTS 或检查网络"
            elif not hint and cs["status"] == "unknown":
                hint = "检查 collector 是否在 SOURCE_TIERS 中正确注册"
            action_link = f" → {hint}" if hint else ""
            alerts.append(f"- {label}: {msg}{action_link}")

    w("## ⚠️ 异常警报")
    w()
    if alerts:
        for a in alerts:
            w(a)
    else:
        w("✅ 全部采集源正常，无异常。")
    w()

    # === 多维度信号聚合 — 值得分析 ===
    w("## 🎯 值得关注的标的")
    w()
    chain_map = _build_code_chain_map()
    signals = _aggregate_signals(dr)

    def _chain_cell(code: str) -> str:
        chains = chain_map.get(code, [])
        if not chains:
            return "—", "—"
        c = chains[0]  # 取第一顺位
        l2 = c["l2"] if c["l2"] and c["l2"] != "-" else ""
        link = f"{c['l1']}>{l2}" if l2 else c["l1"]
        return c["plate"], link

    if signals["deep"] or signals["watch"]:
        w("### 🔬 建议深度分析")
        w()
        w("| 代码 | 名称 | 板块 | 链环节 | 信号分 | 公告 | 研报 | 新闻 | 行业 | 社交 | 催化 | 源 |")
        w("|------|------|------|------|:-----:|:---:|:---:|:---:|:---:|:---:|:---:|:--:|")
        for s in signals["deep"]:
            plate, link = _chain_cell(s["code"])
            w(f"| {s['code']} | {s['name']} | {plate} | {link} | **{s['total']}** | "
              f"{s['dr_count'] or '—'} | {s['rpt_count'] or '—'} | {s['ns_count'] or '—'} | "
              f"{s['ind_count'] or '—'} | {s['social_count'] or '—'} | {s['cat_count'] or '—'} | "
              f"{s['sources']} |")
        w()
        w("### 👀 建议初步关注")
        w()
        w("| 代码 | 名称 | 板块 | 链环节 | 信号分 | 公告 | 研报 | 新闻 | 行业 | 社交 | 催化 | 源 |")
        w("|------|------|------|------|:-----:|:---:|:---:|:---:|:---:|:---:|:---:|:--:|")
        for s in signals["watch"]:
            plate, link = _chain_cell(s["code"])
            w(f"| {s['code']} | {s['name']} | {plate} | {link} | {s['total']} | "
              f"{s['dr_count'] or '—'} | {s['rpt_count'] or '—'} | {s['ns_count'] or '—'} | "
              f"{s['ind_count'] or '—'} | {s['social_count'] or '—'} | {s['cat_count'] or '—'} | "
              f"{s['sources']} |")
        w()
    else:
        w("_本周暂无强信号标的_")
        w()

    # 行业/主题信号热度
    hot = _hot_themes()
    if hot:
        w("### 🔥 信号热度 TOP 主题")
        w()
        w("| 热度 | 主题 | 信号数 | 来源 | 链环节 | 代表标的 |")
        w("|:----:|------|:-----:|------|------|---------|")
        for h in hot[:15]:
            w(f"| {'⭐'*min(h['heat'],3)} | {h['theme']} | {h['count']} | {h['source_list']} | {h.get('chain_tags', '')} | {h['stocks']} |")
        w()

    # === 产业链下钻（逻辑×走势双轨验证） ===
    chain_rows = _chain_heat()
    if chain_rows:
        w("### 🔗 逻辑×走势双轨验证")
        w()
        plates_seen = []
        for ch in chain_rows:
            if ch["plate"] not in plates_seen:
                plates_seen.append(ch["plate"])
                w(f"**{ch['plate']}**  ")
        w()
        w("| 板块 | 层级1 | 层级2 | 标的 | 涨跌 | 逻辑提及 | 走势信号 | 预期差 | 判断 |")
        w("|------|------|------|:--:|:---:|:-------:|:-------:|:-----:|:----:|")
        for ch in chain_rows[:25]:
            l2_display = ch['l2'] if ch['l2'] and ch['l2'] != '-' else '—'
            best_label = f"{ch['top_name']}({ch['top_code']})"
            pct_str = f"+{ch['avg_pct']}%" if ch['avg_pct'] > 0 else f"{ch['avg_pct']}%"
            mention_str = f"{ch['total_mentions']}次/{ch['mention_density']:.0f}密"
            w(f"| {ch['plate']} | {ch['l1']} | {l2_display} | {ch['count']} | {pct_str} | {mention_str} | {ch['trend_signal']} | {ch['expectation_gap']} | **{ch['verdict']}** |")
        w()

    # === 主题生命周期 ===
    lc_rows = _render_theme_lifecycle(w)
    if lc_rows:
        _render_theme_advisor(w, lc_rows, chain_rows)

    # === 双轨选股（FEV + G-Factor） ===
    _render_dual_track(w)

    # === 盘面侧信号（自下而上） ===
    _render_price_signals(w, last_trade)

    # === 本周趋势 ===
    w("## 📈 本周公告深研趋势")
    w()
    daily = dr.get("daily", {})
    if daily:
        max_cnt = max(v[0] for v in daily.values()) or 1
        for d_sorted in sorted(daily.keys()):
            cnt, a60 = daily[d_sorted]
            bar_len = max(1, int(cnt / max_cnt * 20))
            bar = "█" * bar_len
            a60_flag = f"  **{a60}条≥60**" if a60 > 0 else ""
            w(f"- {d_sorted}: {bar} {cnt}条{a60_flag}")
    else:
        w("_暂无数据_")
    w()

    # === 最近提交 ===
    w("## 📜 最近提交")
    w()
    try:
        import subprocess
        result = subprocess.run(
            ["git", "log", "--oneline", "-10"],
            capture_output=True, text=True, timeout=5,
            cwd=str(Path(__file__).parent.parent),
            encoding="utf-8", errors="replace",
        )
        if result.returncode == 0:
            for line in result.stdout.strip().split("\n"):
                if line.strip():
                    w(f"- `{line}`")
        else:
            w("_git 不可用_")
    except Exception:
        w("_git 不可用_")
    w()

    # === 评分规则（自文档化） ===
    w("---")
    w()
    w("## 📐 评分规则")
    w()
    w("### 复合分公式")
    w()
    w("```")
    w("总分 = 公告deep_read分×1.0 + 研报篇数×20 + 新闻信号条数×8 + 行业提及次数×10 + 社交源提及次数×10 + 催化行动分×0.8 + 交叉加分")
    w("```")
    w()
    w("### 各源权重")
    w()
    w("| 源 | 单位 | 权重 | 分值范围 |")
    w("|----|------|:---:|:------:|")
    w("| 公告 deep_read | score(10~85) | ×1.0 | 10~85 |")
    w("| 研报 | 篇数 | ×20 | 0~N×20 |")
    w("| 新闻边际信号 | 条数 | ×8 | 0~N×8 |")
    w("| 行业深研提及 | 次数 | ×10 | 0~N×10 |")
    w("| 社交源提及 | 次数 | ×10 | 0~N×10 |")
    w("| 催化信号 | actionability | ×0.8 | 0~64 |")
    w()
    w("### 交叉源加分")
    w()
    w("| 覆盖源数 | 2 | 3 | 4 | 5 |")
    w("|:-------:|:--:|:--:|:--:|:--:|")
    w("| 加分 | +10 | +20 | +35 | +50 |")
    w()
    w("### 分级阈值")
    w()
    w("| 级别 | 阈值 | 含义 |")
    w("|------|:---:|------|")
    w("| 🔬 深度分析 | ≥80 | 多源强信号或单一源极高强度 |")
    w("| 👀 初步关注 | 20~79 | 1篇研报或2次社交提及即达标 |")
    w()

    # === 底部快速链接 ===
    w("---")
    w()
    w("### 🔗 快速链接")
    w("| 页面 | 路径 |")
    w("|------|------|")
    w("| 复盘报告 | `daily_review/reports/review/review_{date}.md` |")
    w("| 盘前建议 | `daily_review/reports/advice/advice_{date}.md` |")
    w("| 公告深研 | `daily_review/reports/deep_read/` |")
    w("| 个股档案 | `daily_review/reports/research_dossiers/` |")
    w("| 催化筛查 | `daily_review/reports/catalyst/` |")
    w("| 行业研报 | `daily_review/reports/industry/` |")
    w("| 📐 评分体系架构 | `docs/评分体系架构图.md` |")
    w("| Git 历史 | 终端: `git log --oneline` |")
    w("| 会话转录 | `~/.claude/projects/C--Users-daixin-myclaude/*.jsonl` (grep 关键词) |")

    content = "\n".join(L)
    DASHBOARD_PATH.write_text(content, encoding="utf-8")
    return content
